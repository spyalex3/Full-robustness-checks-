import os
import tempfile

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

README_TEXT = [
    ("Sheet", "What it contains / how to read it"),
    ("Run_Info", "Input file, detected date column, variables used, and parameters for this run."),
    (
        "Stationarity",
        "ADF (H0: unit root, non-stationary) and KPSS (H0: stationary) per variable. "
        "If ADF says non-stationary and KPSS agrees, consider differencing before "
        "trusting level regressions, or rely on the Johansen cointegration result instead.",
    ),
    (
        "OLS_Diagnostics",
        "One row per dependent variable (regressed on all other variables, levels, OLS). "
        "Flags: multicollinearity (max VIF > 10), Ramsey RESET (functional form misspecification, "
        "p<0.05 = reject correct specification), Breusch-Godfrey (residual autocorrelation, "
        "p<0.05 = present), Breusch-Pagan/White (heteroskedasticity, p<0.05 = present), "
        "Jarque-Bera (non-normal residuals, p<0.05 = reject normality).",
    ),
    ("OLS_Coefficients", "Coefficient, std error, t-stat, p-value and 95% CI for every regressor in every single-equation model."),
    ("VIF", "Variance Inflation Factor per regressor per equation. VIF > 10 is the conventional multicollinearity warning threshold."),
    ("VAR_LagSelection", "VAR lag order selected by each information criterion (AIC/BIC/HQIC/FPE) and the lag actually used."),
    (
        "Granger_Causality",
        "Pairwise Granger causality from the fitted VAR. H0: 'causing' does NOT Granger-cause 'caused'. "
        "p<0.05 = reject H0 (evidence of predictive causality, not structural causality).",
    ),
    (
        "Johansen_Cointegration",
        "Johansen trace and max-eigenvalue tests for cointegrating relationships among the (assumed I(1)) variables. "
        "'rejects_5pct'=True for rank r means the null of at most r cointegrating vectors is rejected at 5%.",
    ),
    ("IRF_Data", "Numeric orthogonalized (Cholesky) impulse response values: response of each variable to a one-unit shock in each variable, by horizon."),
    ("IRF_Plot", "Plotted orthogonalized impulse response functions with 95% confidence bands. Note: Cholesky ordering follows the column order in your input file."),
    ("Warnings", "Any test that failed to compute (e.g. due to small sample size) and why."),
]


def _autofit(ws, df, max_width=40):
    for i, col in enumerate(df.columns, start=1):
        width = min(max_width, max(12, len(str(col)) + 2, df[col].astype(str).map(len).max() + 2 if len(df) else 12))
        ws.column_dimensions[get_column_letter(i)].width = width


def build_report(
    output_path,
    run_info_df,
    stationarity_df,
    diag_df,
    coef_df,
    vif_frames,
    lag_table_df,
    granger_df,
    johansen_df,
    irf_df,
    irf_png_path,
    warnings,
):
    vif_long = []
    for dep, vdf in vif_frames.items():
        for _, row in vdf.iterrows():
            vif_long.append({"dependent": dep, **row.to_dict()})
    vif_df = pd.DataFrame(vif_long)

    warnings_df = pd.DataFrame({"warning": warnings}) if warnings else pd.DataFrame({"warning": []})
    readme_df = pd.DataFrame(README_TEXT, columns=["Sheet", "Description"])

    sheets = {
        "ReadMe": readme_df,
        "Run_Info": run_info_df,
        "Stationarity": stationarity_df,
        "OLS_Diagnostics": diag_df,
        "OLS_Coefficients": coef_df,
        "VIF": vif_df,
        "VAR_LagSelection": lag_table_df,
        "Granger_Causality": granger_df,
        "Johansen_Cointegration": johansen_df,
        "IRF_Data": irf_df,
        "Warnings": warnings_df,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    for name, df in sheets.items():
        ws = wb[name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        _autofit(ws, df)

    if irf_png_path and os.path.exists(irf_png_path):
        ws = wb.create_sheet("IRF_Plot")
        img = XLImage(irf_png_path)
        ws.add_image(img, "A1")

    wb.save(output_path)
